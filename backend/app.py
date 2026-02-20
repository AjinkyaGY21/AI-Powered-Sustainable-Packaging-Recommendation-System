"""
EcoPackAI Backend
-------------------------------------------------
FIXES APPLIED:
1. SyntaxError in get_materials() — missing closing }) 
2. Cookie config broken on localhost — now runtime-detected
   via @app.before_request (reads request.host each request)
3. Removed unused `from reportlab.pdfgen import canvas`
4. BI dashboard path kept relative via PROJECT_ROOT
"""

from flask import Flask, request, jsonify, session, send_file
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, timedelta
import secrets
import logging
import os
from dotenv import load_dotenv
from io import BytesIO
from pathlib import Path
import sys
import pandas as pd

# PDF imports
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# ==========================================================
# PATH FIX
# ==========================================================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

# ==========================================================
# ML IMPORT
# ==========================================================
try:
    from ml.notebooks.recommendation_engine import (
        generate_recommendations,
        materials_df,
        co2_model,
        cost_model,
        FEATURES_COST,
        FEATURES_CO2
    )
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False

# ==========================================================
# ENV + APP
# ==========================================================
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("APP_SECRET_KEY", secrets.token_urlsafe(32))

# Session lifetime — expires only on logout
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=365)
app.config["SESSION_COOKIE_HTTPONLY"] = True
# NOTE: SameSite / Secure are set per-request below (runtime detection)

CORS(app,
     supports_credentials=True,
     origins=[
         "http://localhost:3000",
         "http://localhost:5500",
         "http://127.0.0.1:5500",
         "http://127.0.0.1:3000",
         "https://ecopackai-web.vercel.app",
         "https://*.vercel.app"
     ],
     allow_headers=["Content-Type"],
     max_age=3600)

# ==========================================================
# FIX 1 — RUNTIME COOKIE POLICY
# localhost/127.0.0.1 → SameSite=Lax, Secure=False  (plain HTTP)
# any other host      → SameSite=None, Secure=True   (HTTPS)
# Zero config needed — just run it anywhere.
# ==========================================================
@app.before_request
def _set_cookie_policy():
    host = request.host.split(":")[0]
    is_local = host in ("localhost", "127.0.0.1")
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"  if is_local else "None"
    app.config["SESSION_COOKIE_SECURE"]   = False   if is_local else True

# Rate limiter (global guard — session-based IP limit is the main one)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "100 per hour"],
    storage_uri="memory://"
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("EcoPackAI")

# ==========================================================
# RATE LIMIT CONFIG (PER CLIENT IP)
# ==========================================================
MAX_RECOMMENDATIONS_PER_WINDOW = 3
RATE_LIMIT_WINDOW_MINUTES = 20

# In-memory store: { ip: {"window_start": datetime, "count": int} }
RATE_LIMITS = {}

# ==========================================================
# MATERIALS DATA
# ==========================================================
MATERIALS_CSV_PATH = PROJECT_ROOT / "data" / "processed" / "clean_materials.csv"
_materials_cache = None

def load_materials_data():
    global _materials_cache
    if _materials_cache is None and MATERIALS_CSV_PATH.exists():
        try:
            _materials_cache = pd.read_csv(MATERIALS_CSV_PATH)
            logger.info(f"Loaded {len(_materials_cache)} materials from CSV")
        except Exception as e:
            logger.error(f"Error loading materials CSV: {e}")
            _materials_cache = pd.DataFrame()
    return _materials_cache if _materials_cache is not None else pd.DataFrame()

# ==========================================================
# SESSION INIT
# (Session stores last shipment/report data only — NOT rate limits)
# ==========================================================
def init_session():
    now = datetime.now()
    if "session_id" not in session:
        session.permanent = True
        session["session_id"] = secrets.token_urlsafe(16)
        session["session_start"] = now.isoformat()
        logger.info(f"New session: {session['session_id']}")
    else:
        session.setdefault("session_start", now.isoformat())

# ==========================================================
# RATE LIMIT CHECK (PER IP, ROLLING WINDOW)
# ==========================================================
def check_recommendation_limit():
    """Returns (allowed, error_msg, used, remaining)"""
    now = datetime.now()
    client_key = get_remote_address()
    record = RATE_LIMITS.get(client_key)

    # Reset window if expired or first visit
    if not record or now - record["window_start"] > timedelta(minutes=RATE_LIMIT_WINDOW_MINUTES):
        record = {"window_start": now, "count": 0}

    if record["count"] >= MAX_RECOMMENDATIONS_PER_WINDOW:
        elapsed = now - record["window_start"]
        remaining_td = timedelta(minutes=RATE_LIMIT_WINDOW_MINUTES) - elapsed
        minutes = max(1, int(remaining_td.total_seconds() / 60))
        RATE_LIMITS[client_key] = record
        msg = (
            f"You've used your quota of {MAX_RECOMMENDATIONS_PER_WINDOW} calls "
            f"per {RATE_LIMIT_WINDOW_MINUTES} minutes. "
            f"Wait and try again in {minutes} minute{'s' if minutes != 1 else ''}."
        )
        return False, msg, record["count"], 0

    record["count"] += 1
    RATE_LIMITS[client_key] = record
    used = record["count"]
    remaining = max(0, MAX_RECOMMENDATIONS_PER_WINDOW - used)
    logger.info(f"Rate limit for {client_key}: {used}/{MAX_RECOMMENDATIONS_PER_WINDOW}")
    return True, None, used, remaining

# ==========================================================
# AUTH STATUS
# ==========================================================
@app.route("/api/auth/status", methods=["GET"])
def auth_status():
    init_session()
    now = datetime.now()
    client_key = get_remote_address()
    record = RATE_LIMITS.get(client_key)

    if not record or now - record["window_start"] > timedelta(minutes=RATE_LIMIT_WINDOW_MINUTES):
        used, remaining = 0, MAX_RECOMMENDATIONS_PER_WINDOW
    else:
        used = record["count"]
        remaining = max(0, MAX_RECOMMENDATIONS_PER_WINDOW - used)

    return jsonify({
        "authenticated": True,
        "user_email": "ecopackai-user@gmail.com",
        "recommendations_used": used,
        "recommendations_remaining": remaining
    })

# ==========================================================
# LOGOUT
# ==========================================================
@app.route("/api/auth/logout", methods=["POST"])
def logout():
    sid = session.get("session_id", "unknown")
    session.clear()
    logger.info(f"Session {sid} logged out")
    init_session()  # Fresh session immediately
    return jsonify({
        "success": True,
        "message": "Logged out successfully",
        "new_session_id": session.get("session_id")
    })

# ==========================================================
# RECOMMEND
# ==========================================================
@app.route("/api/recommend", methods=["POST"])
def recommend():
    allowed, error, used, remaining = check_recommendation_limit()
    if not allowed:
        return jsonify({
            "error": error,
            "limit_reached": True,
            "session_info": {"used": used, "remaining": remaining}
        }), 429

    data = request.get_json() or {}

    required = [
        "Category_item", "Weight_kg", "Fragility",
        "Moisture_Sens", "Distance_km", "Shipping_Mode",
        "Length_cm", "Width_cm", "Height_cm"
    ]
    for field in required:
        if field not in data:
            return jsonify({"error": f"Missing field: {field}"}), 400

    shipment = {
        "Category_item": data["Category_item"],
        "Weight_kg":     float(data["Weight_kg"]),
        "Fragility":     int(data["Fragility"]),
        "Moisture_Sens": bool(data["Moisture_Sens"]),
        "Distance_km":   float(data["Distance_km"]),
        "Shipping_Mode": data["Shipping_Mode"],
        "Length_cm":     float(data["Length_cm"]),
        "Width_cm":      float(data["Width_cm"]),
        "Height_cm":     float(data["Height_cm"]),
    }

    top_k  = int(data.get("top_k", 5))
    sort_by = data.get("sort_by", "Sustainability")

    try:
        if ML_AVAILABLE:
            df = generate_recommendations(
                materials_df, co2_model, cost_model,
                shipment, FEATURES_COST, FEATURES_CO2, top_k, sort_by
            )
            recommendations = df.to_dict("records")
        else:
            logger.warning("ML models not available, returning empty recommendations")
            recommendations = []

        session["last_recommendation"] = recommendations
        session["last_shipment"]        = shipment

        return jsonify({
            "status": "success",
            "recommendations": recommendations,
            "session_info": {"used": used, "remaining": remaining}
        })

    except Exception as e:
        logger.error(f"Error generating recommendations: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==========================================================
# PDF — EcoPackAI header + Shipment Details + Recommendations
# ==========================================================
@app.route("/api/generate-pdf", methods=["POST"])
def generate_pdf():
    data     = session.get("last_recommendation")
    shipment = session.get("last_shipment")

    if not data:
        return jsonify({"error": "Generate recommendation first"}), 400

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        topMargin=0.75*inch, bottomMargin=0.75*inch,
        leftMargin=0.75*inch, rightMargin=0.75*inch
    )
    elements = []
    styles   = getSampleStyleSheet()

    # App title — green
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=20, textColor=colors.HexColor("#10b981"), alignment=1
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=13, textColor=colors.HexColor("#047857")
    )

    elements.append(Paragraph("<b>EcoPackAI</b>", title_style))
    elements.append(Spacer(1, 20))

    # Shipment Details
    if shipment:
        elements.append(Paragraph("<b>Shipment Details:</b>", section_style))
        elements.append(Spacer(1, 8))

        shipment_rows = [
            ["Category",         str(shipment.get("Category_item", "N/A"))],
            ["Weight (kg)",      f"{shipment.get('Weight_kg', 0):.2f}"],
            ["Distance (km)",    f"{shipment.get('Distance_km', 0):.2f}"],
            ["Shipping Mode",    str(shipment.get("Shipping_Mode", "N/A"))],
            ["Fragility",        str(shipment.get("Fragility", "N/A"))],
            ["Moisture Sens.",   "Yes" if shipment.get("Moisture_Sens") else "No"],
            ["Dimensions (cm)",  (f"{shipment.get('Length_cm', 0)} x "
                                  f"{shipment.get('Width_cm', 0)} x "
                                  f"{shipment.get('Height_cm', 0)}")],
        ]

        st = Table(shipment_rows, colWidths=[2*inch, 4*inch])
        st.setStyle(TableStyle([
            ("FONTNAME",  (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",  (0, 0), (-1, -1), 10),
            ("ALIGN",     (0, 0), (0, -1), "LEFT"),
            ("ALIGN",     (1, 0), (1, -1), "CENTER"),
            ("GRID",      (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1),
             [colors.HexColor("#f0fdf4"), colors.white]),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 20))

    # Recommendations table
    elements.append(Paragraph("<b>Recommendations:</b>", section_style))
    elements.append(Spacer(1, 8))

    table_data = [["#", "Material", "Cost ($)", "CO₂ (kg)", "Sustainability"]]
    for i, r in enumerate(data, 1):
        table_data.append([
            str(i),
            str(r.get("Material_Name", "N/A"))[:30],
            f"{r.get('Pred_Cost', 0):.2f}",
            f"{r.get('Pred_CO2', 0):.2f}",
            f"{r.get('Sustainability', 0):.4f}",
        ])

    rec_table = Table(
        table_data,
        colWidths=[0.4*inch, 2.5*inch, 1*inch, 1*inch, 1.3*inch]
    )
    rec_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#10b981")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  10),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE",    (0, 1), (-1, -1), 9),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#f0fdf4"), colors.white]),
    ]))
    elements.append(rec_table)

    # Faint watermark
    def add_watermark(c, doc):
        c.saveState()
        c.setFillColor(colors.Color(0.1, 0.6, 0.4, alpha=0.05))
        c.setFont("Helvetica-Bold", 50)
        c.translate(letter[0] / 2, letter[1] / 2)
        c.rotate(30)
        c.drawCentredString(0, 0, "EcoPackAI")
        c.restoreState()

    doc.build(elements, onFirstPage=add_watermark, onLaterPages=add_watermark)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="EcoPackAI_Report.pdf",
        mimetype="application/pdf"
    )

# ==========================================================
# FIX 2 — MATERIALS: missing closing }) fixed
# ==========================================================
@app.route("/api/materials", methods=["GET"])
def get_materials():
    page      = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 12))

    mat_df = load_materials_data()

    # FIX: was missing closing }) — caused SyntaxError
    if mat_df.empty:
        return jsonify({
            "materials": [],
            "total": 0,
            "page": page,
            "page_size": page_size,
            "has_more": False
        })                        # ← this closing }) was absent in the original

    start_idx = (page - 1) * page_size
    end_idx   = start_idx + page_size

    page_df  = mat_df.iloc[start_idx:end_idx]
    has_more = end_idx < len(mat_df)

    materials_list = page_df.to_dict("records")
    # Sanitise NaN → None so JSON serialisation never breaks
    for m in materials_list:
        for k, v in m.items():
            if isinstance(v, float) and (v != v):  # NaN check
                m[k] = None

    return jsonify({
        "materials": materials_list,
        "total":     len(mat_df),
        "page":      page,
        "page_size": page_size,
        "has_more":  has_more
    })

# ==========================================================
# EXCEL EXPORT
# ==========================================================
@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    data     = session.get("last_recommendation")
    shipment = session.get("last_shipment")

    if not data:
        return jsonify({"error": "Generate recommendation first"}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "EcoPackAI Report"

        # App name header
        ws["A1"] = "EcoPackAI"
        ws["A1"].font      = Font(size=16, bold=True, color="10B981")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:E1")

        # Shipment Details
        ws["A3"] = "Shipment Details:"
        ws["A3"].font = Font(bold=True, size=12)

        shipment_rows = [
            ["Category",        shipment.get("Category_item", "N/A")],
            ["Weight (kg)",     shipment.get("Weight_kg", 0)],
            ["Distance (km)",   shipment.get("Distance_km", 0)],
            ["Shipping Mode",   shipment.get("Shipping_Mode", "N/A")],
            ["Fragility",       shipment.get("Fragility", "N/A")],
            ["Moisture Sens.",  "Yes" if shipment.get("Moisture_Sens") else "No"],
            ["Dimensions (cm)", (f"{shipment.get('Length_cm', 0)} x "
                                 f"{shipment.get('Width_cm', 0)} x "
                                 f"{shipment.get('Height_cm', 0)}")],
        ]

        for idx, row in enumerate(shipment_rows, start=4):
            ws[f"A{idx}"] = row[0]
            ws[f"A{idx}"].font = Font(bold=True)
            ws[f"B{idx}"] = row[1]
            ws[f"B{idx}"].alignment = Alignment(horizontal="center")

        # Recommendations
        ws["A12"] = "Recommendations:"
        ws["A12"].font = Font(bold=True, size=12)

        headers = ["#", "Material", "Cost ($)", "CO₂ (kg)", "Sustainability"]
        for ci, header in enumerate(headers, 1):
            c = ws.cell(row=13, column=ci, value=header)
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="10B981")
            c.alignment = Alignment(horizontal="center")

        for ri, rec in enumerate(data, start=14):
            ws.cell(ri, 1, ri - 13).alignment          = Alignment(horizontal="center")
            ws.cell(ri, 2, rec.get("Material_Name","")).alignment = Alignment(horizontal="center")
            ws.cell(ri, 3, round(rec.get("Pred_Cost", 0), 2)).alignment  = Alignment(horizontal="center")
            ws.cell(ri, 4, round(rec.get("Pred_CO2", 0), 2)).alignment   = Alignment(horizontal="center")
            ws.cell(ri, 5, round(rec.get("Sustainability", 0), 4)).alignment = Alignment(horizontal="center")

        # Auto column widths
        for ci in range(1, ws.max_column + 1):
            max_len = max(
                (len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            as_attachment=True,
            download_name="EcoPackAI_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==========================================================
# BI DASHBOARD
# Path: PROJECT_ROOT/bi/EcoPackAI_BI_Dashboard.html
# Relative to this file (backend/app.py): ../bi/EcoPackAI_BI_Dashboard.html
# ==========================================================
BI_DASHBOARD_HTML = PROJECT_ROOT / "bi" / "EcoPackAI_BI_Dashboard.html"


@app.route("/api/bi-dashboard-available", methods=["GET"])
def bi_dashboard_available():
    return jsonify({"available": BI_DASHBOARD_HTML.exists()})


@app.route("/bi/dashboard", methods=["GET"])
def bi_dashboard():
    if not BI_DASHBOARD_HTML.exists():
        return jsonify({"error": "BI dashboard HTML not found at expected path"}), 404
    return send_file(BI_DASHBOARD_HTML, mimetype="text/html")


# ==========================================================
# HEALTH
# ==========================================================
@app.route("/api/health")
def health():
    host = request.host.split(":")[0]
    return jsonify({
        "status":                        "healthy",
        "ml_available":                  ML_AVAILABLE,
        "environment":                   "local" if host in ("localhost", "127.0.0.1") else "production",
        "max_recommendations_per_window": MAX_RECOMMENDATIONS_PER_WINDOW,
        "rate_limit_window_minutes":     RATE_LIMIT_WINDOW_MINUTES,
    })


# ==========================================================
# RUN
# ==========================================================
if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("🌱 EcoPackAI Backend Starting")
    logger.info(f"   Rate limit : {MAX_RECOMMENDATIONS_PER_WINDOW} per {RATE_LIMIT_WINDOW_MINUTES} min (per IP)")
    logger.info(f"   Session    : expires only on logout")
    logger.info(f"   Cookie     : auto-detected (localhost=Lax, prod=None+Secure)")
    logger.info(f"   ML models  : {ML_AVAILABLE}")
    logger.info(f"   BI dashboard: {BI_DASHBOARD_HTML}")
    logger.info("=" * 60)
    load_materials_data()  # Pre-warm cache
    app.run(debug=True, host="0.0.0.0", port=5000)